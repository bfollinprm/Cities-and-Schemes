"""Convert raw-ingest source files to markdown for staging in .context/converted/.

The output is intermediate: a faithful per-source markdown rendering that we then
hand-split into the knowledge-base folders. Source files in raw-ingest/ are not
modified.

Usage:
    uv run python src/ingest/convert.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import mammoth
import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
RAW_DIR = REPO_ROOT / "raw-ingest" / "cloaks-and-conspiracies"
OUT_DIR = REPO_ROOT / ".context" / "converted"


def convert_docx(src: Path, dest: Path) -> None:
    with src.open("rb") as fh:
        result = mammoth.convert_to_markdown(fh)
    dest.write_text(result.value, encoding="utf-8")
    if result.messages:
        warnings = "\n".join(f"- {m.type}: {m.message}" for m in result.messages)
        warn_path = dest.with_suffix(dest.suffix + ".warnings.txt")
        warn_path.write_text(warnings, encoding="utf-8")


def convert_xlsx(src: Path, dest: Path) -> None:
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [
            ["" if cell is None else str(cell) for cell in row]
            for row in ws.iter_rows(values_only=True)
        ]
        rows = [r for r in rows if any(c.strip() for c in r)]
        if not rows:
            continue
        parts.append(f"## Sheet: {sheet_name}\n")
        width = max(len(r) for r in rows)
        rows = [r + [""] * (width - len(r)) for r in rows]
        header = rows[0]
        body = rows[1:]
        parts.append("| " + " | ".join(_md_cell(c) for c in header) + " |")
        parts.append("| " + " | ".join("---" for _ in header) + " |")
        for r in body:
            parts.append("| " + " | ".join(_md_cell(c) for c in r) + " |")
        parts.append("")
    dest.write_text("\n".join(parts), encoding="utf-8")


def _md_cell(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", " ").strip()


def main() -> int:
    if not RAW_DIR.exists():
        print(f"raw-ingest dir missing: {RAW_DIR}", file=sys.stderr)
        return 1
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for src in sorted(RAW_DIR.iterdir()):
        if src.suffix.lower() == ".docx":
            dest = OUT_DIR / f"{src.stem}.md"
            print(f"docx  {src.name} -> {dest.relative_to(REPO_ROOT)}")
            convert_docx(src, dest)
        elif src.suffix.lower() == ".xlsx":
            dest = OUT_DIR / f"{src.stem}.md"
            print(f"xlsx  {src.name} -> {dest.relative_to(REPO_ROOT)}")
            convert_xlsx(src, dest)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
